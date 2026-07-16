"""
Build a peak-values Excel report from a Toronto Hydro Green Button (ESPI) XML file.

For each monthly billing period the report gives the total energy (kWH) and the peak
demand in kW and kVA, both overall and restricted to on-peak intervals (``*_nop``).

The script is reusable: point it at any similarly formatted Green Button file and it
will derive the billing periods, holidays, and DST-correct local times from the data.

Usage:
    uv run build_peak_values.py [INPUT_XML] [OUTPUT_XLSX]

Defaults:
    INPUT_XML   = TH_Electric_Usage_23-11-2024_to_24-06-2026.XML
    OUTPUT_XLSX = Green_Button_Peak_Values.xlsx

Rules (see Prompt_Green_Button_peak_values.md):
  - Billing period for a month runs from the start of the 24th of the previous month to
    the end of the 23rd of the month; ``Billing_period_ending`` is that 23rd (local time).
  - Off-peak = weekends, Toronto statutory holidays (incl. Civic Holiday), and weekday
    local start times < 07:00 or >= 19:00. On-peak = weekday, non-holiday, hour in 7..18.
  - One row per billing period that contains any data. A period with no on-peak interval
    (e.g. a single-weekend day of data) leaves the ``*_nop`` cells blank.
  - Values are reported in kilo units (kWH, kW, kVA); the library decodes them as Wh/W/VA,
    so we divide by 1000.
  - Rows are in descending order of billing period.
"""

from __future__ import annotations

import datetime as dt
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from zoneinfo import ZoneInfo

import holidays
from holidays.constants import OPTIONAL, PUBLIC
from openpyxl import Workbook
from openpyxl.styles import Font

from greenbutton_objects.parse import parse_feed

LOCAL_TZ = ZoneInfo("America/Toronto")

# ReadingType uom -> logical series name.
UOM_TO_SERIES = {"wattHours": "kwh", "watts": "kw", "voltAmps": "kva"}

COLUMNS = [
    "Billing_period_ending",
    "kWH",
    "Max_kW", "Max_kW_Interval", "Max_kW_Interval_utc",
    "Max_kW_nop", "Max_kW_nop_interval", "Max_kW_nop_interval_utc",
    "Max_kVA", "Max_kVA_Interval", "Max_kVA_Interval_utc",
    "Max_kVA_nop", "Max_kVA_nop_interval", "Max_kVA_nop_interval_utc",
]

DATE_FMT = "%Y-%m-%d"
DATETIME_FMT = "%Y-%m-%d %H:%M"


def billing_period_ending(local_date: dt.date) -> dt.date:
    """Billing period ends on the 23rd; days on/after the 24th roll to next month."""
    year, month = local_date.year, local_date.month
    if local_date.day >= 24:
        month += 1
        if month == 13:
            month, year = 1, year + 1
    return dt.date(year, month, 23)


def build_holiday_set(years: range) -> holidays.HolidayBase:
    """Ontario statutory holidays including the (optional) August Civic Holiday."""
    return holidays.CA(subdiv="ON", years=years, categories=(PUBLIC, OPTIONAL))


def is_off_peak(local_start: dt.datetime, holiday_dates) -> bool:
    """Off-peak if weekend, holiday, or weekday local hour < 7 or >= 19."""
    if local_start.weekday() >= 5:  # Saturday=5, Sunday=6
        return True
    if local_start.date() in holiday_dates:
        return True
    hour = local_start.hour
    return hour < 7 or hour >= 19


@dataclass
class Reading:
    utc_start: dt.datetime
    local_start: dt.datetime
    kwh: float          # kWH for the hour (energy delivered)
    kw: float           # kW (power)
    kva: float          # kVA (apparent power)
    on_peak: bool


@dataclass
class PeriodStats:
    ending: dt.date
    local_days: set = field(default_factory=set)
    kwh_total: float = 0.0
    readings: list = field(default_factory=list)


def load_readings(input_xml: str) -> list[Reading]:
    """Parse the XML and merge the three series into per-hour Reading records."""
    usage_points = parse_feed(input_xml)
    if len(usage_points) != 1:
        print(f"warning: expected 1 UsagePoint, found {len(usage_points)}")

    # Collect {series: {utc_start: value_in_base_units}}. Library values are Wh/W/VA.
    series: dict[str, dict[dt.datetime, float]] = defaultdict(dict)
    for up in usage_points:
        for mr in up.meterReadings:
            name = UOM_TO_SERIES.get(mr.readingType.uom.name)
            if name is None:
                continue
            for ir in mr.intervalReadings:
                series[name][ir.timePeriod.start] = ir.value

    missing = {"kwh", "kw", "kva"} - series.keys()
    if missing:
        raise SystemExit(f"error: input file is missing series: {sorted(missing)}")

    # Timestamps must be identical across the three series.
    timestamps = set(series["kwh"])
    for name in ("kw", "kva"):
        if set(series[name]) != timestamps:
            print(f"warning: {name} series timestamps differ from kwh series")

    holiday_dates = build_holiday_set(
        range(min(timestamps).astimezone(LOCAL_TZ).year,
              max(timestamps).astimezone(LOCAL_TZ).year + 1)
    )
    print("Holidays applied (off-peak):")
    for d in sorted(holiday_dates):
        print(f"  {d}  {holiday_dates[d]}")

    readings: list[Reading] = []
    for utc_start in sorted(timestamps):
        local_start = utc_start.astimezone(LOCAL_TZ)
        readings.append(Reading(
            utc_start=utc_start,
            local_start=local_start,
            kwh=series["kwh"][utc_start] / 1000.0,   # Wh -> kWH
            kw=series["kw"].get(utc_start, 0.0) / 1000.0,   # W  -> kW
            kva=series["kva"].get(utc_start, 0.0) / 1000.0,  # VA -> kVA
            on_peak=not is_off_peak(local_start, holiday_dates),
        ))
    return readings


def group_periods(readings: list[Reading]) -> list[PeriodStats]:
    """Group readings into billing periods (one per period that contains any data)."""
    periods: dict[dt.date, PeriodStats] = {}
    for r in readings:
        ending = billing_period_ending(r.local_start.date())
        ps = periods.setdefault(ending, PeriodStats(ending=ending))
        ps.local_days.add(r.local_start.date())
        ps.kwh_total += r.kwh
        ps.readings.append(r)

    return [periods[k] for k in sorted(periods)]


def first_max(readings: list[Reading], value):
    """Return (max_value, reading) for the first (earliest) interval attaining the max.

    ``readings`` must be in ascending time order. Returns (None, None) if empty.
    """
    best_val = None
    best_reading = None
    for r in readings:
        v = value(r)
        if best_val is None or v > best_val:
            best_val = v
            best_reading = r
    return best_val, best_reading


def row_for_period(ps: PeriodStats) -> dict:
    """Compute the 14-column row for one billing period."""
    row = {c: None for c in COLUMNS}
    row["Billing_period_ending"] = ps.ending.strftime(DATE_FMT)
    row["kWH"] = ps.kwh_total

    on_peak = [r for r in ps.readings if r.on_peak]

    for series, prefix in (("kw", "Max_kW"), ("kva", "Max_kVA")):
        get = (lambda r, s=series: getattr(r, s))

        val, r = first_max(ps.readings, get)
        if r is not None:
            row[prefix] = val
            row[f"{prefix}_Interval"] = r.local_start.strftime(DATETIME_FMT)
            row[f"{prefix}_Interval_utc"] = r.utc_start.strftime(DATETIME_FMT)

        # On-peak-only variant (column names use lowercase "_interval").
        val, r = first_max(on_peak, get)
        if r is not None:
            row[f"{prefix}_nop"] = val
            row[f"{prefix}_nop_interval"] = r.local_start.strftime(DATETIME_FMT)
            row[f"{prefix}_nop_interval_utc"] = r.utc_start.strftime(DATETIME_FMT)

    return row


def write_workbook(rows: list[dict], output_xlsx: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Peak Values"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:  # already sorted descending by caller
        ws.append([row[c] for c in COLUMNS])

    # Reasonable column widths.
    for i, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(name) + 2)
    ws.freeze_panes = "A2"

    wb.save(output_xlsx)


def main() -> None:
    input_xml = sys.argv[1] if len(sys.argv) > 1 else "TH_Electric_Usage_23-11-2024_to_24-06-2026.XML"
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else "Green_Button_Peak_Values.xlsx"

    print(f"Reading {input_xml} ...")
    readings = load_readings(input_xml)
    print(f"  {len(readings)} hourly readings")

    periods = group_periods(readings)
    rows = [row_for_period(ps) for ps in periods]
    rows.sort(key=lambda r: r["Billing_period_ending"], reverse=True)  # descending

    write_workbook(rows, output_xlsx)
    print(f"\nWrote {output_xlsx}: {len(rows)} billing periods x {len(COLUMNS)} columns")
    print(f"  range: {rows[-1]['Billing_period_ending']} .. {rows[0]['Billing_period_ending']}")


if __name__ == "__main__":
    main()
